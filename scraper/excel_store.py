"""
Persistence: appends (or updates) one scrape result as a row in
products.xlsx.
"""

import os
from openpyxl import Workbook, load_workbook

COLUMNS = [
    "Product ID",
    "Product Label",
    "Scraped Title",
    "Per Kg Price",
    "Country of Origin",
    "Supermarket",
    "URL",
    "Timestamp",
]


def save_to_excel(data, filepath="products.xlsx", product_id=None, product_label=None):
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(COLUMNS)  # type: ignore

    today = data["current_time"][:10]

    # --------------------------------------------
    # Check whether today's record already exists
    # --------------------------------------------

    for row in range(2, sheet.max_row + 1):  # type: ignore

        existing_product = sheet.cell(row, 1).value  # type: ignore
        existing_market = sheet.cell(row, 6).value  # type: ignore
        existing_timestamp = sheet.cell(row, 8).value  # type: ignore

        if existing_timestamp:
            existing_date = str(existing_timestamp)[:10]
        else:
            existing_date = ""

        if (
            existing_product == product_id
            and existing_market == data["supermarket"]
            and existing_date == today
        ):
            # Update existing row instead of appending
            sheet.cell(row, 3).value = data["product"]  # type: ignore
            sheet.cell(row, 4).value = data.get("per_kg_price")  # type: ignore
            sheet.cell(row, 5).value = data["country_of_origin"]  # type: ignore
            sheet.cell(row, 7).value = data["url"]  # type: ignore
            sheet.cell(row, 8).value = data["current_time"]  # type: ignore

            wb.save(filepath)
            return

    # --------------------------------------------
    # No row for today -> append new row
    # --------------------------------------------

    sheet.append([  # type: ignore
        product_id,
        product_label,
        data["product"],
        data.get("per_kg_price"),
        data["country_of_origin"],
        data["supermarket"],
        data["url"],
        data["current_time"],
    ])

    wb.save(filepath)
