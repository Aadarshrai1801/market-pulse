"""
One-time migration: copies your existing local data (users.db,
products.xlsx, and a Firebase CSV export) into MongoDB. Safe to re-run -
everything is upserted by its natural key, so running it twice won't
create duplicates.

Usage:
    python migrate_to_mongo.py

Requires MONGODB_URI (and optionally MONGO_DB_NAME) to already be set in
your environment or a local .env file - see README's Environment
variables table. Doesn't touch or delete users.db / products.xlsx / the
Firebase CSV - this only reads them.
"""

import csv
import os
import sqlite3
from datetime import datetime

from openpyxl import load_workbook

from db import get_db, ensure_indexes

USERS_DB_PATH = os.environ.get("USERS_DB_PATH", "users.db")
EXCEL_PATH = os.environ.get("PRODUCTS_XLSX_PATH", "products.xlsx")
FIREBASE_CSV_PATH = os.environ.get("FIREBASE_CSV_PATH", "firebase.csv")

# The Firebase CSV export has a "Product" name and "Retailer" name, not the
# product_id/supermarket slugs price_history is keyed by. Confirmed against
# products.xlsx (the already-migrated data) that product_id is a lowercase/
# underscore slug and supermarket is lowercase/no-spaces - these two maps
# translate the CSV's labels to those same slugs so the migrated rows join
# up with what the live scrapers already write, instead of forking into a
# separate, disconnected series per product. If a future export includes a
# product name not listed here, add it rather than auto-slugifying, so we
# never silently invent an id that doesn't match the app's existing catalog.
FIREBASE_PRODUCT_LABEL_TO_ID = {
    "Cucumber": "cucumber",
    "Garlic": "garlic",
    "Orange Navel": "orange_navel",
    "Orange Valencia": "orange_valencia",
    "Potato": "potato",
    "Red Onion": "red_onion",
    "Tomato": "tomato",
    "Watermelon": "watermelon",
}
FIREBASE_RETAILER_TO_SUPERMARKET = {
    "Barakat": "barakat",
    "Carrefour": "carrefour",
    "Kibsons": "kibsons",
    "LuLu": "lulu",
    "Union Coop": "unioncoop",
}


def migrate_users():
    if not os.path.exists(USERS_DB_PATH):
        print(f"[skip] {USERS_DB_PATH} not found - nothing to migrate.")
        return

    db = get_db()
    conn = sqlite3.connect(USERS_DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT * FROM users").fetchall()
    conn.close()

    max_id = 0
    migrated = 0
    for row in rows:
        user = dict(row)
        user["is_active"] = bool(user["is_active"])

        # sqlite gives us an explicit google_id: None for local accounts.
        # Writing that as $set would store a literal null in Mongo, which
        # defeats the sparse unique index (sparse only skips documents
        # where the field is *missing*, not merely null). So instead of
        # $set-ing the whole dict as-is, $unset google_id when there's no
        # real value - this also cleans up any null left over from a
        # previous (buggy) run of this script.
        update = {"$set": user}
        if not user.get("google_id"):
            user.pop("google_id", None)
            update["$unset"] = {"google_id": ""}

        db.users.update_one({"id": user["id"]}, update, upsert=True)
        max_id = max(max_id, user["id"])
        migrated += 1

    if max_id:
        # Keep the auto-increment counter ahead of the highest migrated id,
        # so the next user created through the app doesn't collide.
        db.counters.update_one({"_id": "users"}, {"$max": {"seq": max_id}}, upsert=True)

    print(f"[users] migrated {migrated} account(s) from {USERS_DB_PATH}.")


def migrate_products():
    if not os.path.exists(EXCEL_PATH):
        print(f"[skip] {EXCEL_PATH} not found - nothing to migrate.")
        return

    db = get_db()
    wb = load_workbook(EXCEL_PATH, read_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True)) #type: ignore
    wb.close()

    if len(rows) < 2:
        print(f"[products] {EXCEL_PATH} has no data rows.")
        return

    header = rows[0]
    wanted = [
        "Product ID", "Product Label", "Scraped Title", "Per Kg Price",
        "Country of Origin", "Supermarket", "URL", "Timestamp",
    ]
    idx = {name: header.index(name) for name in wanted if name in header}

    migrated = 0
    skipped = 0
    for row in rows[1:]:
        timestamp = row[idx["Timestamp"]] if "Timestamp" in idx else None
        product_id = row[idx["Product ID"]] if "Product ID" in idx else None
        supermarket = row[idx["Supermarket"]] if "Supermarket" in idx else None

        if not timestamp or not product_id or not supermarket:
            skipped += 1
            continue

        date = str(timestamp)[:10]
        record = {
            "product_id": product_id,
            "product_label": row[idx["Product Label"]] if "Product Label" in idx else None,
            "scraped_title": row[idx["Scraped Title"]] if "Scraped Title" in idx else None,
            "per_kg_price": row[idx["Per Kg Price"]] if "Per Kg Price" in idx else None,
            "country_of_origin": row[idx["Country of Origin"]] if "Country of Origin" in idx else None,
            "supermarket": supermarket,
            "url": row[idx["URL"]] if "URL" in idx else None,
            "timestamp": timestamp,
            "date": date,
        }
        db.price_history.update_one(
            {"product_id": product_id, "supermarket": supermarket, "date": date},
            {"$set": record},
            upsert=True,
        )
        migrated += 1

    print(f"[products] migrated {migrated} price record(s) from {EXCEL_PATH} ({skipped} row(s) skipped - missing key fields).")


def migrate_firebase_history():
    if not os.path.exists(FIREBASE_CSV_PATH):
        print(f"[skip] {FIREBASE_CSV_PATH} not found - nothing to migrate.")
        return

    db = get_db()

    migrated = 0
    skipped = 0

    with open(FIREBASE_CSV_PATH, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    for row in rows:
        label = row["Product"].strip()
        retailer = row["Retailer"].strip()

        product_id = FIREBASE_PRODUCT_LABEL_TO_ID.get(label)
        supermarket = FIREBASE_RETAILER_TO_SUPERMARKET.get(retailer)

        if not product_id or not supermarket:
            print(
                f"[skip] unrecognized Product/Retailer: {label!r} / {retailer!r} - "
                f"add it to the mapping dicts near the top of this script if it's legitimate."
            )
            skipped += 1
            continue

        try:
            price = float(row["Price (AED)"])
        except (TypeError, ValueError):
            print(f"[skip] unparseable price {row['Price (AED)']!r} for {label} / {retailer}")
            skipped += 1
            continue

        try:
            # "12/06/2026, 9:44:33 AM" -> "2026-06-12 09:44:33", to match the
            # timestamp format already used by mongo_store.py / products.xlsx.
            timestamp = datetime.strptime(
                row["Fetched At"].strip(), "%d/%m/%Y, %I:%M:%S %p"
            ).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            print(f"[skip] unparseable 'Fetched At' {row['Fetched At']!r} for {label} / {retailer}")
            skipped += 1
            continue

        date = row["Date"].strip()

        record = {
            "product_id": product_id,
            "product_label": label,
            "scraped_title": label,  # CSV has no separate scraped title
            "per_kg_price": f"AED {price:.2f}/kg",  # CSV Unit is always "per kg"
            "country_of_origin": row["Origin Country"].strip(),
            "supermarket": supermarket,
            "url": None,  # CSV has no product URL
            "timestamp": timestamp,
            "date": date,
        }

        db.price_history.update_one(
            {"product_id": product_id, "supermarket": supermarket, "date": date},
            {"$set": record},
            upsert=True,
        )
        migrated += 1

    print(f"[firebase] migrated {migrated} price record(s) from {FIREBASE_CSV_PATH} ({skipped} row(s) skipped).")


if __name__ == "__main__":
    print(f"Connecting to MongoDB (database: {os.environ.get('MONGO_DB_NAME', 'intellicrop')})...")
    ensure_indexes()
    migrate_users()
    # Order matters here: both migrate_products() and migrate_firebase_history()
    # upsert into price_history on the same key (product_id, supermarket, date),
    # so whichever runs LAST wins for any date both sources have data for.
    # Firebase is treated as the more authoritative/recent source, so it must
    # run after migrate_products() - don't reorder these two calls.
    migrate_products()
    migrate_firebase_history()
    print("Done. users.db, products.xlsx, and the Firebase CSV were left untouched - this only reads them.")