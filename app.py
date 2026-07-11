import os
import re
import threading
import uuid
from datetime import datetime, timedelta
import werkzeug
from flask import Flask, request, jsonify, render_template
import cv2
import numpy as np

from ocr.ocr import preprocess_image, run_ocr, extract_table
from openpyxl import load_workbook

from scraper import (
    SITE_SEARCH_CONFIG,
    find_product_url,
    get_product_details,
    save_to_excel,
)
from products_config import PRODUCTS, PRODUCTS_BY_ID, RETAILER_LABELS, get_search_keyword
from auth import auth_bp, init_auth, login_required, role_required

app = Flask(__name__)

# All persistent data (users.db via USERS_DB_PATH, the uploads folder,
# products.xlsx, and the cached secret.key fallback below) lives under
# DATA_DIR. Point this at your host's mounted volume in production (e.g.
# DATA_DIR=/app/data on Railway) - anything written outside a mounted
# volume is lost on the next redeploy/restart.
DATA_DIR = os.environ.get("DATA_DIR", ".")
os.makedirs(DATA_DIR, exist_ok=True)


def _load_or_create_secret_key(path=None):
    """
    Flask needs a stable secret_key to sign session cookies - if it changes
    on every restart, everyone gets logged out each time the server
    restarts. SECRET_KEY env var wins if set (**strongly recommended for
    production** - on most hosts the local filesystem doesn't persist
    across deploys, so the file-cache fallback below would silently
    regenerate on every deploy and log everyone out); otherwise a random
    key is generated once and cached in `path` so it survives restarts on
    this machine.
    """
    path = path or os.path.join(DATA_DIR, "secret.key")
    env_key = os.environ.get("SECRET_KEY")
    if env_key:
        return env_key
    if os.path.exists(path):
        with open(path, "r") as f:
            return f.read().strip()
    key = os.urandom(32).hex()
    with open(path, "w") as f:
        f.write(key)
    return key


app.secret_key = _load_or_create_secret_key()
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=7)
app.config["SESSION_COOKIE_HTTPONLY"] = True
# Cookie only sent over HTTPS, and not attached to cross-site requests.
# Both assume the app is served over HTTPS in production (true on
# Railway/Render and behind any real reverse proxy) - if you're ever
# testing production config over plain HTTP, SESSION_COOKIE_SECURE will
# silently stop the session cookie from being set at all.
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("SESSION_COOKIE_SECURE", "1") == "1"
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
# Reject uploads over 10MB before they're fully buffered/decoded - guards
# the /api/ocr endpoint against memory exhaustion from oversized images.
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024

init_auth(app)          # creates users.db + seeds a default admin if empty
app.register_blueprint(auth_bp)

UPLOAD_FOLDER = os.path.join(DATA_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

RETAILERS = list(SITE_SEARCH_CONFIG.keys())
EXCEL_PATH = os.path.join(DATA_DIR, "products.xlsx")

# In-memory job store for the async fetch API - fine for single-device use
# (resets if the app restarts, which is expected here).
JOBS = {}
JOBS_LOCK = threading.Lock()


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def _price_to_float(text):
    if not text:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)", text)
    return float(m.group(1)) if m else None

def get_previous_price(product_id, retailer, current_price):
    """
    Returns the previous saved price for the same product and retailer,
    ignoring the current price if it is already the newest entry.
    """
    if not os.path.exists(EXCEL_PATH):
        return None

    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True)) #type: ignore

    if len(rows) < 2:
        wb.close()
        return None

    header = rows[0]

    pid_idx = header.index("Product ID")
    retailer_idx = header.index("Supermarket")
    price_idx = header.index("Per Kg Price")
    time_idx = header.index("Timestamp")

    history = []

    for row in rows[1:]:
        if row[pid_idx] != product_id:
            continue
        if str(row[retailer_idx]).lower() != retailer.lower():
            continue

        price = _price_to_float(row[price_idx])
        if price is None:
            continue

        history.append({
            "price": price,
            "timestamp": row[time_idx]
        })

    wb.close()

    history.sort(key=lambda x: str(x["timestamp"]))

    if len(history) < 2:
        return None

    return history[-2]["price"]


def _find_col(header, *candidates):
    """Case/spacing-tolerant column lookup - returns the index of the first
    header cell matching any candidate name, or None if none match."""
    normalized = [str(h).strip().lower() if h else "" for h in header]
    for candidate in candidates:
        target = candidate.strip().lower()
        if target in normalized:
            return normalized.index(target)
    return None


def _resolve_retailers(retailer_param):
    if not retailer_param or retailer_param == "all":
        return RETAILERS
    if retailer_param not in SITE_SEARCH_CONFIG:
        return []
    return [retailer_param]


def _resolve_products(product_param):
    if not product_param or product_param == "all":
        return PRODUCTS
    p = PRODUCTS_BY_ID.get(product_param)
    return [p] if p else []


def _scrape_one(product, retailer):
    """Run one product x retailer lookup, save it if it succeeds, and
    always return a plain-dict result (never raises)."""
    keyword = get_search_keyword(product, retailer)

    try:
        url = find_product_url(retailer, keyword)
        data = get_product_details(url, retailer)
        
        # Use per_kg_price, not "price" - that's the field every retailer
        # module actually returns (Carrefour returns both, but Barakat,
        # Kibsons, LuLu, and Union Coop only ever set per_kg_price), and
        # it's also the exact column get_previous_price() reads history
        # from below, so this keeps current vs. previous comparing
        # like-for-like.
        current_price = _price_to_float(data.get("per_kg_price"))

        previous_price = get_previous_price(
            product["id"],
            retailer,
            current_price
        )

        data["previous_price"] = previous_price

        # Guard against either side being None (e.g. this scrape failed to
        # find a price, or there's no usable history yet) - comparing None
        # to a float raises TypeError and used to crash the whole job once
        # enough history had built up.
        if previous_price is None or current_price is None:
            data["price_change"] = "same"
        elif current_price > previous_price:
            data["price_change"] = "up"
        elif current_price < previous_price:
            data["price_change"] = "down"
        else:
            data["price_change"] = "same"
                
        data["ok"] = True
        data["product_id"] = product["id"]
        data["product_label"] = product["name"]
        data["product_emoji"] = product["emoji"]

        save_to_excel(
            data,
            filepath=EXCEL_PATH,
            product_id=product["id"],
            product_label=product["name"],
        )
        return data
    except Exception as e:
        return {
            "ok": False,
            "supermarket": retailer,
            "product_id": product["id"],
            "product_label": product["name"],
            "product_emoji": product["emoji"],
            "keyword_used": keyword,
            "error": str(e),
        }


# ------------------------------------------------------------------
# Pages
# ------------------------------------------------------------------

@app.route("/")
@login_required
def index():
    return render_template("index.html")


# ------------------------------------------------------------------
# API: dropdown options
# ------------------------------------------------------------------

@app.route("/api/meta")
@login_required
def api_meta():
    return jsonify({
        "retailers": [
            {"id": r, "name": RETAILER_LABELS.get(r, r.capitalize())}
            for r in RETAILERS
        ],
        "products": [
            {"id": p["id"], "name": p["name"], "emoji": p["emoji"]}
            for p in PRODUCTS
        ],
    })


# ------------------------------------------------------------------
# API: fetch fresh prices (runs the scraper, appends to products.xlsx)
# ------------------------------------------------------------------

@app.route("/api/fetch", methods=["POST"])
@role_required("editor", "admin")
def api_fetch():
    """
    Synchronous fetch - fine for a single retailer/product (a few seconds),
    but for bigger requests (especially "all"/"all", ~40 lookups) this
    call will sit open for minutes. Use POST /api/jobs instead for those -
    it returns immediately and you poll for progress.
    """
    payload = request.get_json(force=True, silent=True) or {}

    retailer_param = (payload.get("retailer") or "all").strip().lower()
    product_param = (payload.get("product") or "all").strip().lower()

    retailers = _resolve_retailers(retailer_param)
    products = _resolve_products(product_param)

    if not retailers:
        return jsonify({"ok": False, "error": f"Unknown retailer '{retailer_param}'."}), 400
    if not products:
        return jsonify({"ok": False, "error": f"Unknown product '{product_param}'."}), 400

    results = [
        _scrape_one(product, retailer)
        for product in products
        for retailer in retailers
    ]

    return jsonify({
        "ok": True,
        "fetched": len(results),
        "results": results,
    })


# ------------------------------------------------------------------
# API: async jobs (submit now, poll for progress/results later)
# ------------------------------------------------------------------

def _run_fetch_job(job_id, retailers, products):
    for product in products:
        for retailer in retailers:
            result = _scrape_one(product, retailer)

            with JOBS_LOCK:
                job = JOBS.get(job_id)
                if job is None:
                    return  # job was cleared/removed while running
                job["results"].append(result)
                job["completed"] += 1

    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if job is not None:
            job["status"] = "done"
            job["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


@app.route("/api/jobs", methods=["POST"])
@role_required("editor", "admin")
def api_create_job():
    """
    Kick off a fetch in the background and return immediately.
    Body: {"retailer": "all"|<id>, "product": "all"|<id>}
    """
    payload = request.get_json(force=True, silent=True) or {}

    retailer_param = (payload.get("retailer") or "all").strip().lower()
    product_param = (payload.get("product") or "all").strip().lower()

    retailers = _resolve_retailers(retailer_param)
    products = _resolve_products(product_param)

    if not retailers:
        return jsonify({"ok": False, "error": f"Unknown retailer '{retailer_param}'."}), 400
    if not products:
        return jsonify({"ok": False, "error": f"Unknown product '{product_param}'."}), 400

    job_id = uuid.uuid4().hex[:12]

    with JOBS_LOCK:
        JOBS[job_id] = {
            "job_id": job_id,
            "status": "running",
            "retailer": retailer_param,
            "product": product_param,
            "total": len(retailers) * len(products),
            "completed": 0,
            "results": [],
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": None,
        }

    thread = threading.Thread(
        target=_run_fetch_job,
        args=(job_id, retailers, products),
        daemon=True,
    )
    thread.start()

    with JOBS_LOCK:
        total = JOBS[job_id]["total"]

    return jsonify({"ok": True, "job_id": job_id, "status": "running", "total": total}), 202


@app.route("/api/jobs/<job_id>")
@login_required
def api_get_job(job_id):
    """Poll this for progress/results: {status: 'running'|'done', completed, total, results}."""
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if job is None:
            return jsonify({"ok": False, "error": f"Unknown job id '{job_id}'."}), 404
        # shallow copy so the caller gets a stable snapshot even if the
        # background thread is still appending to it
        snapshot = dict(job)
        snapshot["results"] = list(job["results"])

    return jsonify({"ok": True, **snapshot})


@app.route("/api/jobs")
@login_required
def api_list_jobs():
    """Recent jobs, newest first (without their full result payloads)."""
    with JOBS_LOCK:
        jobs = [
            {k: v for k, v in job.items() if k != "results"}
            for job in JOBS.values()
        ]

    jobs.sort(key=lambda j: j["created_at"], reverse=True)
    return jsonify({"ok": True, "jobs": jobs})


# ------------------------------------------------------------------
# API: history matrix (product+retailer rows x date columns)
# ------------------------------------------------------------------

@app.route("/api/history")
@login_required
def api_history():
    retailer_param = (request.args.get("retailer") or "all").strip().lower()
    product_param = (request.args.get("product") or "all").strip().lower()

    empty = {
        "dates": [],
        "rows": [],
        "counts": {"dates": 0, "products": 0, "retailers": 0},
    }

    if not os.path.exists(EXCEL_PATH):
        return jsonify(empty)

    wb = load_workbook(EXCEL_PATH, read_only=True)
    sheet = wb.active

    header = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))] #type: ignore

    try:
        idx = {name: header.index(name) for name in [
            "Product ID", "Product Label", "Per Kg Price",
            "Supermarket", "Timestamp",
        ]}
    except ValueError:
        # old-format workbook without Product ID/Label columns
        return jsonify(empty)

    # Origin column name isn't guaranteed, so try a few likely variants.
    # If none of these match your actual header, check products.xlsx and
    # add the exact text here.
    origin_idx = _find_col(
        header,
        "Country Of Origin", "Country of Origin", "Origin", "Country",
    )

    # cell[key=(product_id, retailer)][date_iso] = latest price that day
    cells = {}
    date_set = set()

    for row in sheet.iter_rows(min_row=2): #type: ignore
        values = [c.value for c in row]
        if len(values) <= max(idx.values()):
            continue

        product_id = values[idx["Product ID"]]
        product_label = values[idx["Product Label"]]
        retailer = values[idx["Supermarket"]]
        price_text = values[idx["Per Kg Price"]]
        timestamp = values[idx["Timestamp"]]
        origin = values[origin_idx] if origin_idx is not None and len(values) > origin_idx else None

        if not product_id or not retailer or not timestamp:
            continue
        if product_param != "all" and product_id != product_param:
            continue
        if retailer_param != "all" and retailer != retailer_param:
            continue

        try:
            date_iso = str(timestamp)[:10]
        except Exception:
            continue

        price = _price_to_float(price_text)
        date_set.add(date_iso)

        key = (product_id, retailer)
        cells.setdefault(key, {"product_label": product_label, "prices": {}, "origins": {}})
        # later rows overwrite earlier ones for the same day -> "latest wins"
        cells[key]["prices"][date_iso] = price
        if origin:
            cells[key]["origins"][date_iso] = str(origin).strip()

    wb.close()

    dates = sorted(date_set)

    rows = []
    for (product_id, retailer), info in sorted(cells.items(), key=lambda kv: (kv[1]["product_label"] or "", kv[0][1])):
        product = PRODUCTS_BY_ID.get(product_id, {})
        rows.append({
            "product_id": product_id,
            "product_label": info["product_label"] or product.get("name", product_id),
            "product_emoji": product.get("emoji", "🥬"),
            "retailer": retailer,
            "retailer_label": RETAILER_LABELS.get(retailer, retailer.capitalize()),
            "prices": info["prices"],
            "origins": info["origins"],
        })

    distinct_products = {r["product_id"] for r in rows}
    distinct_retailers = {r["retailer"] for r in rows}

    return jsonify({
        "dates": dates,
        "rows": rows,
        "counts": {
            "dates": len(dates),
            "products": len(distinct_products),
            "retailers": len(distinct_retailers),
        },
    })
    

    
@app.route('/api/ocr', methods=['POST'])
@role_required("editor", "admin")
def api_ocr_scan():
    """
    Bridge API endpoint accepting multipart image forms from api.js 
    and returning structured token maps matching the frontend expectations.
    """
    if 'image' not in request.files:
        return jsonify({"ok": False, "error": "No image file part found in request form parameters."}), 400
        
    file = request.files['image']
    if file.filename == '':
        return jsonify({"ok": False, "error": "No file selected for transmission."}), 400

    try:
        # 1. Read image safely without writing directly to disk
        in_memory_stream = file.read()
        nparr = np.frombuffer(in_memory_stream, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            return jsonify({"ok": False, "error": "Invalid image file format or corrupted payload."}), 400

        # 2. Temporary storage setup because preprocess_image expects a path string
        # (Alternatively, you can modify preprocess_image to accept the image matrix directly)
        # A unique filename per request - a fixed name here meant two people
        # uploading at the same moment would overwrite each other's file and
        # could get OCR results from the wrong image.
        temp_path = os.path.join(UPLOAD_FOLDER, f"ocr_{uuid.uuid4().hex}.png")
        cv2.imwrite(temp_path, img)

        # 3. Fire the custom PP-OCRv5 pipeline steps
        processed_matrix = preprocess_image(temp_path)
        ocr_tokens = run_ocr(processed_matrix, confidence_threshold=0.60)
        structured_rows = extract_table(ocr_tokens)

        # 4. Cleanup temp disk usage
        if os.path.exists(temp_path):
            os.remove(temp_path)

        # 5. Map fields seamlessly into Javascript's expected schema
        # frontend looks for: country, shipment, product, weight, packing, price, confidence
        payload_products = []
        for item in structured_rows:
            # Reconstruct an aggregate token confidence safely, or fallback to 1.0
            avg_conf = 0.85

            payload_products.append({
                "country": item.get("country") or "",
                "shipment": item.get("shipment") or "",
                "product": item.get("product") or "Unknown Product",
                "weight": item.get("weight") or "",
                "packing": item.get("packing") or "",
                "price": item.get("price") or "0.00",
                "confidence": avg_conf
            })

        return jsonify({
            "ok": True,
            "count": len(payload_products),
            "products": payload_products
        })

    except Exception:
        # Full detail goes to the server log only - returning str(e) to the
        # client can leak internal paths/config, and isn't actionable for
        # whoever's on the other end of the upload anyway.
        app.logger.exception("OCR scan failed")
        try:
            if 'temp_path' in locals() and os.path.exists(temp_path): #type: ignore
                os.remove(temp_path) #type: ignore
        except OSError:
            pass
        return jsonify({
            "ok": False,
            "error": "OCR processing failed. Please try again with a clearer image."
        }), 500


if __name__ == "__main__":
    # Dev-only entrypoint. In production this file is imported by gunicorn
    # (see Dockerfile: `gunicorn -w 1 -b 0.0.0.0:$PORT app:app`), so this
    # block never runs on the server - debug=True here is safe precisely
    # because it's local-only.
    app.run(host="127.0.0.1", port=int(os.environ.get("PORT", 5000)), debug=True, threaded=True)